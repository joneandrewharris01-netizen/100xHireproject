"""Load LinkedIn leads from an xlsx file (header row required)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def load(path: Path | str) -> list[dict]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({h: (str(v).strip() if v is not None else "") for h, v in zip(headers, r)})
    return out
